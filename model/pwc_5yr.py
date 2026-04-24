"""Stage 6 — Integrated 5-Year Fiscal Model for Prince William County, VA.

Purpose
-------
Integrate the five preceding stages into a reproducible FY27-FY31 fiscal
model with exactly three scenarios:

Pre-Cancellation Digital Gateway
    Digital Gateway (DG) proceeds on the pre-ruling schedule. The published
    FY26-FY30 Adopted 5-year forecast (PWC-REV-FY26-30 p.2) is held intact.
    Non-DC revenue follows Stage 4 BASE.  Committed expenditure = Stage 5
    BASE (schools transfer computed from the Pre-Cancellation Digital Gateway
    eligible-revenue base through `schools_transfer()` from Stage 5).  The
    FY26 FY26-FY30 Adopted forecast was published BEFORE the April 2026
    non-appeal of the Oak Valley ruling; it therefore embeds intact overlay
    reliability.

CAPEX Spillover
    - Canceled Pageland parcels (Stage 2) are removed from DC revenue on
      the Stage 2 phasing schedule (50/200/450/750/1,050 MW for FY27-FY31
      at $231K/MW/yr derived in research/canceled_projects.md §4.3).
    - New (non-Pageland) overlay CAPEX is deterred at Stage 3's synthesized
      trajectory: 35/25/18/12/10% year-1..year-5 reduction on a ~500
      MW/yr expected new-build pipeline priced at $235K/MW/yr
      (data/spillover_parameters.csv new_capex_reduction_pct_year*).
    - Accelerated write-down: contingent overlay land assessed value takes
      a 25% haircut (assessed_value_haircut_contingent_overlay), applied
      probability-weighted by the 40% impairment-disclosure probability
      (impairment_disclosure_probability_24mo).  The base CAPEX Spillover
      case applies the expected-value markdown; the sensitivity layer
      (exposed on the sheet `impairment_sensitivity`) applies the full
      100% markdown in FY28 and propagates it forward.
    - Non-DC revenue = Stage 4 LOW plus Stage 3b residential regulatory-
      spillover overlay (peer-county-grounded; Loudoun 2025 residential-
      permit differential of -53% YoY was the primary benchmark). Overlay
      reduces new-construction AV by Year-N permit_units_reduction and
      applies Year-N residential_av_growth_drag additively to residential
      AV growth. Commercial and industrial channels have NULL point
      estimates per Stage 3b's finance-office-honest call; ranges carried
      for sensitivity but point case does not move those lines vs.
      Pre-Cancellation Digital Gateway.
    - Expenditure = Stage 5 BASE + Stage 6b credit-spread incremental
      debt service. Stage 3b §e prices a 1-notch Moody's action (Aaa ->
      Aa1) at +12 bps on cumulative new-money GO issuance FY27+; Year-5
      point estimate ~$390K/yr, full range $165K-$1.12M.
    - State SOQ receives the LCI 2-year lag effect per Stage 4
      (pwc_lci_drift LOW scenario): assessed-value haircut compounds PWC's
      local-share index, reducing state aid from FY29+.

Partial Recovery
    - Canceled Pageland projects stay canceled.  The Va. Ct. App.
      affirmance is final (Oak Valley, Mar 31 2026); the BoCS declined to
      appeal on Apr 15 2026; the specific zoning ordinances were held VOID
      AB INITIO and cannot be retroactively restored by a policy pivot.
    - Assume the county issues a clarifying ordinance or settles with the
      Oak Valley plaintiffs within 12-24 months.  Half of the
      CAPEX Spillover deterred CAPEX returns starting FY29, subject to a
      30% confidence discount (i.e. 70% of baseline volume).
    - Non-DC revenue = Stage 4 BASE (the shock mostly relents).
    - Expenditure = Stage 5 BASE + HALF credit-spread effect. Stage 3b
      §e.5 notes that rating recovery typically lags policy reversal by
      12-24+ months; a partial-recovery path therefore does NOT immediately
      undo the rating action. The half-effect captures this asymmetry.

Model outputs (per scenario x FY)
---------------------------------
1. Total general-fund revenue (DC + non-DC) disaggregated by source.
2. Schools transfer via `schools_transfer(eligible_revenue_base)` from
   Stage 5.  The eligible base is the "Total General Revenues" line from
   the 2013 PWC-Schools Revenue Sharing Agreement: it INCLUDES data-center
   taxes on the real-property + business-tangible + BPOL + fees lines that
   already count toward the GR split, and EXCLUDES Agency Revenue, PPTRA
   fixed reimbursement, federal/state categorical aid, fire levy,
   stormwater, solid waste fees, and CSA pass-through (per the docstring
   in model/expenditures.py::schools_transfer).
3. Total committed expenditure (Stage 5 BASE; scenario-sensitive only
   through the schools-transfer line, which re-solves on each scenario's
   eligible base rather than the published forecast base).
4. Surplus/deficit = revenue - expenditure.
5. Required RE tax rate per $100 assessed value to close any deficit.
   Using scenario-specific assessed value (which under CAPEX Spillover is
   haircut on the DC-zoned land portion).  Current FY26 rate is $0.906/$100
   (baseline CSV; PWC-REV-FY26-30 p.2).
6. Debt-service-to-revenue ratio with the PSFM 10% cap.
7. Reserve trajectory applied to FY26 Unassigned GF balance of $134.7M
   (baseline CSV; PWC-ACFR-FY25 p.35) vs. the 7.5% policy floor.
8. Impairment sensitivity layer: CAPEX Spillover at full markdown vs.
   expected markdown.

Reproducibility
---------------
`python3 model/pwc_5yr.py` regenerates every output CSV and the .xlsx.  No
network access; reads only:
    data/pwc_baseline.csv        (Stage 1 anchor)
    data/canceled_projects.csv   (Stage 2 DG phasing)
    data/spillover_parameters.csv (Stage 3 elasticities)
    data/non_dc_revenue_projection.csv  (Stage 4 BASE/LOW/HIGH)
    data/revenue_driver_assumptions.csv (Stage 4 assumption table)
    data/expenditure_path.csv    (Stage 5 base path)
    data/expenditure_assumptions.csv (Stage 5 assumption table)
    data/debt_service_schedule.csv   (Stage 5 contractual debt service)

and imports:
    model.revenue_drivers        (Stage 4)
    model.expenditures           (Stage 5 incl. schools_transfer())

Source keys
-----------
All source keys resolve to citations/stage1-5.bib (inherited).  New
Stage 6 citations, if any, go in citations/stage6.bib.
"""

from __future__ import annotations

import csv
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd

# Stage 4 and Stage 5 modules (imported to demonstrate integration and to
# source schools_transfer() directly from Stage 5).
import importlib.util

REPO = Path(__file__).resolve().parent.parent
DATA = REPO / "data"
MODEL = REPO / "model"

# --- Import Stage 5 expenditures module (avoid package machinery) ---
_spec = importlib.util.spec_from_file_location(
    "_stage5_expenditures", MODEL / "expenditures.py"
)
_expenditures = importlib.util.module_from_spec(_spec)
# Register in sys.modules BEFORE exec so that @dataclass can resolve the
# module namespace (dataclasses.py consults sys.modules[cls.__module__]).
sys.modules["_stage5_expenditures"] = _expenditures
_spec.loader.exec_module(_expenditures)  # type: ignore
schools_transfer = _expenditures.schools_transfer
county_share = _expenditures.county_share
check_debt_capacity = _expenditures.check_debt_capacity
credit_spread_scenario_c = _expenditures.credit_spread_scenario_c

# Scenario identifiers used as internal keys, CSV column values, and xlsx
# sheet names. Long forms are the authoritative names; SHORT_NAMES supplies
# a space-tight variant for figure labels and xlsx tab names.
SCENARIO_PRE_DG: str = "Pre-Cancellation Digital Gateway"
SCENARIO_SPILLOVER: str = "CAPEX Spillover"
SCENARIO_PARTIAL: str = "Partial Recovery"

SHORT_NAMES: Dict[str, str] = {
    SCENARIO_PRE_DG: "Pre-Cancel DG",
    SCENARIO_SPILLOVER: "CAPEX Spillover",
    SCENARIO_PARTIAL: "Partial Recovery",
}

# Stage 6b — scenario factor applied to the credit-spread overlay.
# Pre-Cancellation Digital Gateway: 0 (Aaa preserved under pre-ruling world).
# CAPEX Spillover: 1.0 (full 1-notch action per Stage 3b).
# Partial Recovery: 0.5 (rating recovery lags policy reversal 12-24+ months
#   per Stage 3b §e.5; partial effect retained).
CREDIT_SPREAD_FACTOR_BY_SCENARIO: Dict[str, float] = {
    SCENARIO_PRE_DG: 0.0,
    SCENARIO_SPILLOVER: 1.0,
    SCENARIO_PARTIAL: 0.5,
}

# ---------------------------------------------------------------------------
# Inputs, anchored constants and forecast horizon.
# ---------------------------------------------------------------------------
FISCAL_YEARS: Tuple[int, ...] = (2027, 2028, 2029, 2030, 2031)
SCENARIOS: Tuple[str, ...] = (SCENARIO_PRE_DG, SCENARIO_SPILLOVER, SCENARIO_PARTIAL)

# FY26 baseline anchors -- extracted from pwc_baseline.csv.
FY26_RE_TAX_RATE_PER_100: float = 0.906  # $/ $100 of AV (PWC-REV-FY26-30 p.2).
FY26_TOTAL_ASSESSED_VALUE: float = 137_561_804_000.0  # FY25 ACFR p.309 Table 14.
FY26_DC_TOTAL_ASSESSED_VALUE: float = 21_700_000_000.0  # PWC-RE-2025 p.23.
FY26_UNASSIGNED_GF_BALANCE: float = 134_724_000.0  # FY25 audited, PWC-ACFR-FY25 p.35.
FY26_DC_REVENUE_STABILIZATION_RESERVE: float = 12_073_000.0  # PWC-ACFR-FY25 p.101.
FY26_GF_REVENUE_TOTAL: float = 1_978_688_390.0  # PWC-BUD-FY26-REV p.55.
GR_SPLIT_BASE_FY26: float = 1_732_673_500.0  # PWC-REV-FY26-30 p.2.
# Agency Revenue and Fire Levy are earmarked / dept-specific revenues that
# offset matching expenditures.  Including them on the revenue side makes
# the surplus/deficit calc internally consistent with the Stage 5 expenditure
# envelope which contains Fire Levy operating + department operations
# underwritten by agency fees.  They are NEITHER in the 57.23% split base
# (schools_transfer excludes them per expenditures.py docstring).
FY26_AGENCY_REVENUE: float = 240_300_000.0  # PWC-BUD-FY26-REV p.63.
FY26_FIRE_LEVY_REVENUE: float = 93_448_379.0  # PWC-BUD-FY26-EXP; matches Fire Levy operating.
# Growth rates for these lines -- agency revenue grows with population/service
# load (use 3%/yr base); Fire Levy revenue grows with Fire Levy tax base
# which tracks residential AV growth and the approved CIP station build-out.
AGENCY_REVENUE_GROWTH_BY_SCENARIO: Dict[str, float] = {
    SCENARIO_PRE_DG: 0.03,
    SCENARIO_SPILLOVER: 0.01,
    SCENARIO_PARTIAL: 0.02,
}
FIRE_LEVY_REVENUE_GROWTH_BY_SCENARIO: Dict[str, float] = {
    SCENARIO_PRE_DG: 0.06,
    SCENARIO_SPILLOVER: 0.035,
    SCENARIO_PARTIAL: 0.05,
}
RE_POLICY_UNASSIGNED_FLOOR_FRACTION: float = 0.075  # PSFM 1.02.
DEBT_SERVICE_CAP_FRACTION: float = 0.10  # PSFM 5.02(d).

# Published FY27-FY30 GR split-base (pre-ruling baseline).  Source:
# PWC-REV-FY26-30 p.2 / baseline CSV.  FY31 extrapolated at FY29-FY30
# growth (see expenditures.py::project_schools_transfer).
PUBLISHED_GR_BASE: Dict[int, float] = {
    2027: 1_807_905_700.0,
    2028: 1_889_734_014.0,
    2029: 1_974_751_404.0,
    2030: 2_063_873_884.0,
}
# FY29->FY30 growth rate from the adopted plan = 4.51%.  Used for FY31.
_GR_FY30_FY29_GROWTH = PUBLISHED_GR_BASE[2030] / PUBLISHED_GR_BASE[2029] - 1.0
PUBLISHED_GR_BASE[2031] = PUBLISHED_GR_BASE[2030] * (1 + _GR_FY30_FY29_GROWTH)

# Personal property vehicle rate FY26 is in the non-DC stream; the DC
# stream is CE&P on business-tangible + DC real-property.
# Per Stage 1 FY26 DC revenue derived estimate ~$350M (baseline_notes.md
# §3), we use the same figure as the DC-revenue anchor in FY26 terms.
# This is labeled "derived" in the baseline CSV; revisions as the TY2025
# data-center report is published (Fall 2026) will flow through this row.
FY26_DC_REVENUE_DERIVED: float = 350_000_000.0  # PWC-DCR-TY24 p.26 derived.

# Stage 2 / Stage 3 phasing for canceled Digital Gateway projects.  MW in
# service (IT load) per research/canceled_projects.md §4.3 Table.
CANCELED_DG_PHASING_MW: Dict[int, float] = {
    2027: 50.0,
    2028: 200.0,
    2029: 450.0,
    2030: 750.0,
    2031: 1_050.0,
}
# Stage 3 / Stage 2 derived per-MW yield at stabilization.
DC_REVENUE_PER_MW_POINT: float = 235_000.0  # pwt2023dgvote @ Stage 3 CSV.

# Stage 3 spillover parameters.
NEW_CAPEX_REDUCTION_BY_YEAR: Dict[int, float] = {
    2027: 0.35,
    2028: 0.25,
    2029: 0.18,
    2030: 0.12,
    2031: 0.10,
}
# Expected volume of *new* (non-Pageland) overlay CAPEX absent the ruling,
# expressed in MW of IT load per fiscal year.  Calibration: Stage 3 §e.2
# notes "~60M sqft / ~6,000 MW of announced-but-not-yet-rezoned capacity
# countywide."  That is the 5-year pipeline; per-year expected siting
# decisions for NEW capacity (i.e. rezoning or SUP approvals) average on
# the order of 500 MW/yr.  Held flat FY27-FY31.
NEW_OVERLAY_CAPEX_PIPELINE_MW_PER_YEAR: float = 500.0
# These MW, once sited, reach *stabilized* revenue (full yield) only after
# ~3-4 years of construction.  CAPEX Spillover's deterrence therefore hits
# fiscal impact on a *lagged* basis: siting year t -> meaningful revenue
# years t+3..t+5 at partial yield.  For the 5-year window we adopt a
# simplifying convention: in year t, deterred MW = reduction% x annual
# pipeline and they forego 1/5 of their stabilized yield in year t (a
# linear ramp proxy).  This is the approach used in research/
# spillover_evidence.md §e.2: "compounded across the 5-year window,
# undiscounted cumulative lost fiscal-impact ~$280M point estimate."
DETERRED_FIRST_YEAR_YIELD_FRACTION: float = 0.20

IMPAIRMENT_DISCLOSURE_PROBABILITY_24MO: float = 0.40  # Stage 3 CSV.
ASSESSED_VALUE_HAIRCUT_CONTINGENT_OVERLAY: float = 0.25  # Stage 3 CSV.
# Contingent-overlay AV: Stage 3 cites $1.312B (pwc_vacant_dc_zoned_land_2023)
# as the PWC TY2023 assessed value of vacant DC-zoned land.  Scale up
# modestly to FY26 base (+25% to reflect TY2024-25 appreciation on vacant
# DC-zoned parcels per PWC-RE-2025) and treat as the at-risk base.
VACANT_DC_ZONED_AV_FY26: float = 1_640_000_000.0  # $1.312B x 1.25.

# Stage 4 LCI drift (LOW scenario) applied to state SOQ from FY29+.
# See revenue_driver_assumptions.csv state_aid/pwc_lci_drift,low = +1%/yr.
# We apply this as an incremental reduction to the state_aid_gf_direct
# line which was modeled in Stage 4 under the LOW scenario with this
# drift *already included* -- so CAPEX Spillover's state_aid_gf_direct =
# the LOW Stage 4 value (no double-count).  We make this explicit and guard.

# Partial Recovery profile.  From Stage 3 §e.4 the judgment was year-2
# deterrence halves; year-3+ steady residual 5%.  Brief re-frames for the
# current task: "Half of CAPEX Spillover's deterred CAPEX returns starting
# FY29 with a 30% confidence discount (70% of baseline volume)."  This
# means: for FY29-FY31 we add back 50% of CAPEX Spillover's MW deterrent
# at 70% of nominal yield.  FY27-FY28 match CAPEX Spillover in deterrence.
PARTIAL_RECOVERY_RETURN_SHARE: float = 0.50  # Half of deterred CAPEX returns.
PARTIAL_RECOVERY_CONFIDENCE_DISCOUNT: float = 0.30  # At 70% of nominal yield.

# ---------------------------------------------------------------------------
# Data loaders (read-only, no network).
# ---------------------------------------------------------------------------


def _read_csv_skip_comments(path: Path) -> pd.DataFrame:
    with path.open() as fh:
        data = [line for line in fh if not line.lstrip().startswith("#")]
    from io import StringIO
    return pd.read_csv(StringIO("".join(data)))


def load_baseline() -> pd.DataFrame:
    return _read_csv_skip_comments(DATA / "pwc_baseline.csv")


def load_canceled_projects() -> pd.DataFrame:
    return _read_csv_skip_comments(DATA / "canceled_projects.csv")


def load_spillover_parameters() -> pd.DataFrame:
    return _read_csv_skip_comments(DATA / "spillover_parameters.csv")


def load_non_dc_revenue() -> pd.DataFrame:
    # Stage 4 output CSV (no leading comment lines but we defend anyway).
    return _read_csv_skip_comments(DATA / "non_dc_revenue_projection.csv")


def load_expenditure_path() -> pd.DataFrame:
    return _read_csv_skip_comments(DATA / "expenditure_path.csv")


def load_debt_schedule_totals() -> Dict[int, float]:
    totals: Dict[int, float] = {}
    with (DATA / "debt_service_schedule.csv").open() as fh:
        reader = csv.DictReader(
            line for line in fh if not line.lstrip().startswith("#")
        )
        for row in reader:
            if row.get("issuance_name") == "GRAND_TOTAL":
                fy = int(row["fiscal_year"].replace("FY", ""))
                totals[fy] = float(row["total"])
    return totals


# ---------------------------------------------------------------------------
# Revenue scenario builders.
# ---------------------------------------------------------------------------


@dataclass
class RevenueLineRow:
    scenario: str
    fiscal_year: int
    revenue_source: str
    amount_usd: float
    in_split_base: bool  # True if in 57.23% Schools-transfer eligible base.
    notes: str = ""


# Stage 4 lines that are in the 57.23% split base (per
# schools_transfer() docstring and expenditure_assumptions.csv
# schools_transfer/eligible_revenue_base_definition row).
STAGE4_LINES_IN_SPLIT_BASE = {
    "residential_real_estate",
    "non_dc_commercial_real_estate",
    "personal_property_vehicles",
    "non_dc_business_tangible",
    "local_sales_tax",
    "bpol_tax",
    "food_and_beverage_tax",
    "consumer_utility_tax",
    "communications_sales_tax",
    "motor_vehicle_license",
    "recordation_tax",
    "transient_occupancy_tax",
    "cigarette_tax",
    "investment_income",
}
STAGE4_LINES_OUT_OF_SPLIT_BASE = {
    "state_aid_gf_direct",  # Categorical state aid.
    "federal_pilt",         # Federal categorical.
    "pptra_fixed_reimbursement",  # Fixed Commonwealth reimbursement.
}


def _non_dc_rows_for_scenario(non_dc_df: pd.DataFrame, stage4_scenario: str) -> List[RevenueLineRow]:
    rows: List[RevenueLineRow] = []
    df = non_dc_df[non_dc_df["scenario"] == stage4_scenario]
    for _, r in df.iterrows():
        line = r["revenue_line"]
        in_base = line in STAGE4_LINES_IN_SPLIT_BASE
        rows.append(
            RevenueLineRow(
                scenario="",  # set by caller
                fiscal_year=int(r["fiscal_year"]),
                revenue_source=line,
                amount_usd=float(r["amount_usd"]),
                in_split_base=in_base,
                notes=f"Stage 4 {stage4_scenario} scenario",
            )
        )
    return rows


def _published_dc_implicit(fy: int, non_dc_base_split: float) -> float:
    """DC component implicit in the Pre-Cancellation Digital Gateway
    (published) GR forecast.

    Pre-Cancellation Digital Gateway = published GR base - Stage 4 BASE
    split-base components. This is the DC revenue the pre-ruling plan
    assumed would arrive. Any negative result is clamped to 0 (shouldn't
    happen in practice).
    """
    dc = PUBLISHED_GR_BASE[fy] - non_dc_base_split
    return max(dc, 0.0)


def _spillover_dc_revenue(
    fy: int, non_dc_base_split: float
) -> Tuple[float, Dict[str, float]]:
    """CAPEX Spillover DC revenue: Pre-Cancellation Digital Gateway DC
    minus canceled DG phasing minus new-CAPEX deterrence minus
    expected-value accelerated write-down on contingent overlay assessed
    value.

    Returns (total_dc_revenue, component_dict) for audit.
    """
    pre_dg_dc = _published_dc_implicit(fy, non_dc_base_split)

    # 1) Canceled Pageland projects: MW-phasing x per-MW yield (Stage 2 §4.3).
    #    These MW were INCLUDED in the Pre-Cancellation Digital Gateway
    #    forecast (DG was expected to deliver them); CAPEX Spillover
    #    removes them.
    canceled_mw = CANCELED_DG_PHASING_MW[fy]
    canceled_revenue_loss = canceled_mw * DC_REVENUE_PER_MW_POINT

    # 2) New-CAPEX deterrence on non-Pageland overlay.  At reduction% of
    #    the annual pipeline (500 MW/yr), at first-year partial yield.
    year_index = fy - 2027  # 0..4
    reduction_pct = NEW_CAPEX_REDUCTION_BY_YEAR[fy]
    # Cumulative deterred MW through year N is sum of reductions 2027..fy.
    deterred_mw_cumulative = sum(
        NEW_CAPEX_REDUCTION_BY_YEAR[y] * NEW_OVERLAY_CAPEX_PIPELINE_MW_PER_YEAR
        for y in range(2027, fy + 1)
    )
    # First-year-yield proxy: cumulative deterred MW x 20% of stabilized
    # yield.  Assumes the missed CAPEX would have been ramping to partial
    # revenue by the fiscal year in question.  See docstring on
    # DETERRED_FIRST_YEAR_YIELD_FRACTION.
    deterrence_revenue_loss = (
        deterred_mw_cumulative
        * DC_REVENUE_PER_MW_POINT
        * DETERRED_FIRST_YEAR_YIELD_FRACTION
    )

    # 3) Accelerated write-down on contingent overlay land (expected value).
    #    Expected markdown = haircut% x probability of impairment disclosure.
    expected_markdown_av = (
        VACANT_DC_ZONED_AV_FY26
        * ASSESSED_VALUE_HAIRCUT_CONTINGENT_OVERLAY
        * IMPAIRMENT_DISCLOSURE_PROBABILITY_24MO
    )
    # Lost RE tax revenue at the current FY26 rate.  Applied from FY27+
    # because re-assessment following the Apr 2026 non-appeal lands on
    # the TY2026 landbook (for FY27 revenue recognition).
    writedown_revenue_loss = expected_markdown_av * (FY26_RE_TAX_RATE_PER_100 / 100.0)

    dc_total = (
        pre_dg_dc
        - canceled_revenue_loss
        - deterrence_revenue_loss
        - writedown_revenue_loss
    )
    dc_total = max(dc_total, 0.0)

    return dc_total, {
        "pre_dg_dc_implicit": pre_dg_dc,
        "canceled_pageland_loss": canceled_revenue_loss,
        "new_capex_deterrence_loss": deterrence_revenue_loss,
        "accelerated_writedown_re_loss_expected": writedown_revenue_loss,
    }


def _partial_recovery_dc_revenue(
    fy: int, non_dc_base_split: float
) -> Tuple[float, Dict[str, float]]:
    """Partial Recovery DC revenue: Pageland stays canceled.  From FY29+,
    half of CAPEX Spillover's deterred CAPEX returns at 70% nominal yield.
    Impairment probability is unchanged in year-1 and year-2 (can't un-ring
    the bell); write-down reverses proportionally from FY29+ as clarity
    returns."""
    pre_dg_dc = _published_dc_implicit(fy, non_dc_base_split)
    canceled_mw = CANCELED_DG_PHASING_MW[fy]
    canceled_revenue_loss = canceled_mw * DC_REVENUE_PER_MW_POINT

    # Determine deterred MW in CAPEX Spillover terms.
    deterred_mw_cumulative_spill = sum(
        NEW_CAPEX_REDUCTION_BY_YEAR[y] * NEW_OVERLAY_CAPEX_PIPELINE_MW_PER_YEAR
        for y in range(2027, fy + 1)
    )
    # For FY27-FY28: Partial Recovery matches CAPEX Spillover deterrence.
    # For FY29-FY31: half returns at 70% nominal yield.
    if fy <= 2028:
        deterrence_revenue_loss = (
            deterred_mw_cumulative_spill
            * DC_REVENUE_PER_MW_POINT
            * DETERRED_FIRST_YEAR_YIELD_FRACTION
        )
    else:
        # Retained deterrence = 50% x CAPEX Spillover deterrence.
        retained_deterred_mw = deterred_mw_cumulative_spill * (1 - PARTIAL_RECOVERY_RETURN_SHARE)
        # Returning MW contribute at 70% nominal (30% confidence discount).
        returning_mw = deterred_mw_cumulative_spill * PARTIAL_RECOVERY_RETURN_SHARE
        returning_contribution_loss = returning_mw * DC_REVENUE_PER_MW_POINT * (
            DETERRED_FIRST_YEAR_YIELD_FRACTION
        ) * PARTIAL_RECOVERY_CONFIDENCE_DISCOUNT  # the 30% net-lost from returnees
        retained_loss = (
            retained_deterred_mw * DC_REVENUE_PER_MW_POINT * DETERRED_FIRST_YEAR_YIELD_FRACTION
        )
        deterrence_revenue_loss = retained_loss + returning_contribution_loss

    # Write-down: 15% impairment probability (Stage 3 partial_recovery_*)
    # applied at the same 25% haircut.
    partial_prob = 0.15
    expected_markdown_av = (
        VACANT_DC_ZONED_AV_FY26 * ASSESSED_VALUE_HAIRCUT_CONTINGENT_OVERLAY * partial_prob
    )
    writedown_revenue_loss = expected_markdown_av * (FY26_RE_TAX_RATE_PER_100 / 100.0)

    dc_total = (
        pre_dg_dc
        - canceled_revenue_loss
        - deterrence_revenue_loss
        - writedown_revenue_loss
    )
    dc_total = max(dc_total, 0.0)

    return dc_total, {
        "pre_dg_dc_implicit": pre_dg_dc,
        "canceled_pageland_loss": canceled_revenue_loss,
        "new_capex_deterrence_loss": deterrence_revenue_loss,
        "accelerated_writedown_re_loss_expected": writedown_revenue_loss,
    }


def _earmarked_revenue_rows(scenario: str) -> List[RevenueLineRow]:
    """Agency Revenue and Fire Levy revenue.  Earmarked; NOT in 57.23%
    split base.  Grown from FY26 base at scenario-specific rates.
    """
    rows: List[RevenueLineRow] = []
    agency_g = AGENCY_REVENUE_GROWTH_BY_SCENARIO[scenario]
    firelevy_g = FIRE_LEVY_REVENUE_GROWTH_BY_SCENARIO[scenario]
    for fy in FISCAL_YEARS:
        years = fy - 2026
        agency = FY26_AGENCY_REVENUE * ((1 + agency_g) ** years)
        firelevy = FY26_FIRE_LEVY_REVENUE * ((1 + firelevy_g) ** years)
        rows.append(
            RevenueLineRow(
                scenario=scenario,
                fiscal_year=fy,
                revenue_source="agency_revenue",
                amount_usd=agency,
                in_split_base=False,
                notes=(
                    f"FY26 ${FY26_AGENCY_REVENUE:,.0f} grown {agency_g*100:.1f}%/yr. "
                    "Agency fees (recordation, clerk, parks/rec, EMS billing, etc.)."
                ),
            )
        )
        rows.append(
            RevenueLineRow(
                scenario=scenario,
                fiscal_year=fy,
                revenue_source="fire_levy_revenue",
                amount_usd=firelevy,
                in_split_base=False,
                notes=(
                    f"FY26 ${FY26_FIRE_LEVY_REVENUE:,.0f} grown {firelevy_g*100:.1f}%/yr. "
                    "Fire Levy tax; earmarked to Fire Levy operating fund."
                ),
            )
        )
    return rows


def build_revenue_rows() -> pd.DataFrame:
    """Build full revenue detail: rows (scenario x FY x source)."""
    non_dc = load_non_dc_revenue()

    non_dc_base = non_dc[non_dc["scenario"] == "base"]
    non_dc_low = non_dc[non_dc["scenario"] == "low"]
    # Stage 6b — CAPEX Spillover uses LOW + Stage 3b residential regulatory-
    # spillover overlay. revenue_drivers.build_full_projection() emits
    # this fourth series as 'low_with_spillover'. Fall back to pure LOW
    # if the overlay is absent (e.g. first run before Stage 3b updates).
    non_dc_low_spill = non_dc[non_dc["scenario"] == "low_with_spillover"]
    if non_dc_low_spill.empty:
        non_dc_low_spill = non_dc_low

    rows: List[RevenueLineRow] = []

    # Earmarked revenue -- Agency fees + Fire Levy -- one entry per scenario x FY.
    for scenario in SCENARIOS:
        rows.extend(_earmarked_revenue_rows(scenario))

    # ---- Pre-Cancellation Digital Gateway: non-DC = Stage 4 BASE;
    #      DC = pre-DG implicit ----
    for fy in FISCAL_YEARS:
        # Add each non-DC line (from Stage 4 BASE).
        for _, r in non_dc_base[non_dc_base["fiscal_year"] == fy].iterrows():
            line = r["revenue_line"]
            in_base = line in STAGE4_LINES_IN_SPLIT_BASE
            rows.append(
                RevenueLineRow(
                    scenario=SCENARIO_PRE_DG,
                    fiscal_year=fy,
                    revenue_source=line,
                    amount_usd=float(r["amount_usd"]),
                    in_split_base=in_base,
                    notes="Stage 4 BASE (non-DC)",
                )
            )
        # Compute the non-DC split-base total for this FY (Stage 4 BASE).
        non_dc_base_split = float(
            non_dc_base[
                (non_dc_base["fiscal_year"] == fy)
                & (non_dc_base["revenue_line"].isin(STAGE4_LINES_IN_SPLIT_BASE))
            ]["amount_usd"].sum()
        )
        dc_pre = _published_dc_implicit(fy, non_dc_base_split)
        rows.append(
            RevenueLineRow(
                scenario=SCENARIO_PRE_DG,
                fiscal_year=fy,
                revenue_source="data_center_tax_revenue",
                amount_usd=dc_pre,
                in_split_base=True,
                notes=(
                    "Pre-Cancellation Digital Gateway DC = published GR forecast "
                    "- Stage 4 BASE split components. Pre-ruling DG pipeline intact."
                ),
            )
        )

    # ---- CAPEX Spillover: non-DC = Stage 4 LOW + Stage 3b residential
    #      overlay; DC shocked. Stage 3b overlay is peer-county-grounded
    #      (Loudoun 2025 -53% residential permit differential after
    #      by-right DC elimination). ----
    for fy in FISCAL_YEARS:
        for _, r in non_dc_low_spill[non_dc_low_spill["fiscal_year"] == fy].iterrows():
            line = r["revenue_line"]
            in_base = line in STAGE4_LINES_IN_SPLIT_BASE
            rows.append(
                RevenueLineRow(
                    scenario=SCENARIO_SPILLOVER,
                    fiscal_year=fy,
                    revenue_source=line,
                    amount_usd=float(r["amount_usd"]),
                    in_split_base=in_base,
                    notes=(
                        "Stage 4 LOW + Stage 3b residential-spillover overlay "
                        "(peer-county grounded; non-DC; federal contraction "
                        "travels with CAPEX Spillover)"
                    ),
                )
            )
        non_dc_low_split = float(
            non_dc_low_spill[
                (non_dc_low_spill["fiscal_year"] == fy)
                & (non_dc_low_spill["revenue_line"].isin(STAGE4_LINES_IN_SPLIT_BASE))
            ]["amount_usd"].sum()
        )
        # CAPEX Spillover DC subtracts from the Pre-Cancellation Digital
        # Gateway's *published* DC pipeline; we use non_dc_BASE split for
        # that subtraction (i.e. the implicit DC in the pre-ruling forecast
        # was computed against BASE non-DC).
        non_dc_base_split = float(
            non_dc_base[
                (non_dc_base["fiscal_year"] == fy)
                & (non_dc_base["revenue_line"].isin(STAGE4_LINES_IN_SPLIT_BASE))
            ]["amount_usd"].sum()
        )
        dc_spill, comp = _spillover_dc_revenue(fy, non_dc_base_split)
        rows.append(
            RevenueLineRow(
                scenario=SCENARIO_SPILLOVER,
                fiscal_year=fy,
                revenue_source="data_center_tax_revenue",
                amount_usd=dc_spill,
                in_split_base=True,
                notes=(
                    f"CAPEX Spillover DC = Pre-Cancel DG DC ({comp['pre_dg_dc_implicit']/1e6:.1f}M)"
                    f" - canceled Pageland ({comp['canceled_pageland_loss']/1e6:.1f}M)"
                    f" - deterrence ({comp['new_capex_deterrence_loss']/1e6:.1f}M)"
                    f" - write-down expected-value ({comp['accelerated_writedown_re_loss_expected']/1e6:.2f}M)"
                ),
            )
        )

    # ---- Partial Recovery: non-DC = Stage 4 BASE; DC = Partial Recovery ----
    for fy in FISCAL_YEARS:
        for _, r in non_dc_base[non_dc_base["fiscal_year"] == fy].iterrows():
            line = r["revenue_line"]
            in_base = line in STAGE4_LINES_IN_SPLIT_BASE
            rows.append(
                RevenueLineRow(
                    scenario=SCENARIO_PARTIAL,
                    fiscal_year=fy,
                    revenue_source=line,
                    amount_usd=float(r["amount_usd"]),
                    in_split_base=in_base,
                    notes="Stage 4 BASE (non-DC; county reverses course)",
                )
            )
        non_dc_base_split = float(
            non_dc_base[
                (non_dc_base["fiscal_year"] == fy)
                & (non_dc_base["revenue_line"].isin(STAGE4_LINES_IN_SPLIT_BASE))
            ]["amount_usd"].sum()
        )
        dc_partial, comp = _partial_recovery_dc_revenue(fy, non_dc_base_split)
        rows.append(
            RevenueLineRow(
                scenario=SCENARIO_PARTIAL,
                fiscal_year=fy,
                revenue_source="data_center_tax_revenue",
                amount_usd=dc_partial,
                in_split_base=True,
                notes=(
                    f"Partial Recovery DC = Pre-Cancel DG DC ({comp['pre_dg_dc_implicit']/1e6:.1f}M)"
                    f" - Pageland ({comp['canceled_pageland_loss']/1e6:.1f}M)"
                    f" - partial deterrence ({comp['new_capex_deterrence_loss']/1e6:.1f}M)"
                    f" - partial write-down ({comp['accelerated_writedown_re_loss_expected']/1e6:.2f}M)"
                ),
            )
        )

    # Convert to DataFrame.
    df = pd.DataFrame(
        [
            {
                "scenario": r.scenario,
                "fiscal_year": r.fiscal_year,
                "revenue_source": r.revenue_source,
                "amount_usd": r.amount_usd,
                "in_split_base": r.in_split_base,
                "notes": r.notes,
            }
            for r in rows
        ]
    )
    return df


def revenue_totals_and_split_base(revenue_df: pd.DataFrame) -> pd.DataFrame:
    """Return per-scenario, per-FY: total revenue and split-base eligible."""
    rows = []
    for (scenario, fy), g in revenue_df.groupby(["scenario", "fiscal_year"]):
        total = g["amount_usd"].sum()
        split_base = g[g["in_split_base"]]["amount_usd"].sum()
        rows.append(
            {
                "scenario": scenario,
                "fiscal_year": fy,
                "total_revenue": total,
                "split_base_eligible": split_base,
            }
        )
    return pd.DataFrame(rows).sort_values(["scenario", "fiscal_year"]).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Expenditure scenario builder.
# ---------------------------------------------------------------------------


def build_expenditure_rows(revenue_totals: pd.DataFrame) -> pd.DataFrame:
    """Build per-scenario committed expenditure.

    Stage 5 BASE is the backbone for all three scenarios (expenditure
    commitments are contractual / already-approved).  The Schools transfer
    line is the one scenario-sensitive component and is re-solved via
    `schools_transfer(eligible_revenue_base)` on each scenario's split base
    -- NOT recomputed from scratch.  Pension/OPEB rows are INFORMATIONAL
    only (basis='embedded') and do not sum into expenditure.
    """
    exp = load_expenditure_path()
    exp_base = exp[exp["scenario"] == "base"].copy()

    rows = []
    for scenario in SCENARIOS:
        for fy in FISCAL_YEARS:
            fy_str = f"FY{fy}"
            year_rows = exp_base[exp_base["fiscal_year"] == fy_str]
            for _, r in year_rows.iterrows():
                line = r["line_item"]
                block = r["block"]
                basis = r["basis"]
                amount = float(r["amount_usd"])
                if block == "schools_transfer":
                    # Replace with scenario-specific transfer.
                    base_row = revenue_totals[
                        (revenue_totals["scenario"] == scenario)
                        & (revenue_totals["fiscal_year"] == fy)
                    ]
                    split_base = float(base_row["split_base_eligible"].iloc[0])
                    amount = schools_transfer(split_base)
                    note = (
                        f"schools_transfer(split_base={split_base:,.0f}) "
                        f"= 57.23% applied to scenario {scenario} eligible base"
                    )
                else:
                    note = str(r.get("notes", ""))
                rows.append(
                    {
                        "scenario": scenario,
                        "fiscal_year": fy,
                        "block": block,
                        "line_item": line,
                        "amount_usd": amount,
                        "basis": basis,
                        "source_key": r.get("source_key", ""),
                        "notes": note,
                    }
                )

            # Stage 6b — add credit-spread incremental debt service row.
            # Zero for Pre-Cancellation Digital Gateway; full for CAPEX
            # Spillover; half for Partial Recovery. See Stage 3b §e for
            # derivation. Counted as 'programmed' (new-money tranches are
            # on the approved CIP plan). Source: Moody's methodology +
            # PWC Moody's Sept 2025 credit opinion.
            factor = CREDIT_SPREAD_FACTOR_BY_SCENARIO.get(scenario, 0.0)
            cs_amount = credit_spread_scenario_c(fy) * factor
            rows.append(
                {
                    "scenario": scenario,
                    "fiscal_year": fy,
                    "block": "debt_service",
                    "line_item": "spillover_credit_spread_ds",
                    "amount_usd": cs_amount,
                    "basis": "programmed",
                    "source_key": "PWC_MOODYS_OPINION_SEPT2025+PWC-BUD-FY26-CIP",
                    "notes": (
                        f"Stage 6b credit-spread overlay; scenario {scenario} "
                        f"factor {factor:.2f}. 1-notch Moody's action "
                        f"(Aaa->Aa1, +12 bps point, range 5-30) on cumulative "
                        "new-money GO issuance FY27+. Zero for Pre-Cancellation "
                        "Digital Gateway; full for CAPEX Spillover; half for "
                        "Partial Recovery (rating recovery lags policy reversal "
                        "12-24+ mo)."
                    ),
                }
            )
    return pd.DataFrame(rows)


def expenditure_totals(exp_df: pd.DataFrame) -> pd.DataFrame:
    """Sum committed expenditure by scenario x FY, excluding embedded rows."""
    non_embedded = exp_df[exp_df["basis"] != "embedded"]
    totals = (
        non_embedded.groupby(["scenario", "fiscal_year"])["amount_usd"]
        .sum()
        .reset_index()
        .rename(columns={"amount_usd": "total_committed_expenditure"})
    )
    return totals


# ---------------------------------------------------------------------------
# Scenario metrics: required RE rate, debt ratio, reserve trajectory.
# ---------------------------------------------------------------------------


def scenario_assessed_value(scenario: str, fy: int) -> float:
    """Return scenario-specific total assessed value.

    Pre-Cancellation Digital Gateway: grow FY26 AV at 5%/yr (matching the
    implicit AV growth in the published forecast). CAPEX Spillover:
    subtract the full (not expected-value) vacant DC-zoned haircut from
    the FY26 base, then grow the remainder at 2%/yr (residential + non-DC
    commercial per Stage 4 LOW). Partial Recovery: same as CAPEX Spillover
    for FY27-28; recovers half the haircut from FY29.
    """
    years_from_fy26 = fy - 2026
    if scenario == SCENARIO_PRE_DG:
        return FY26_TOTAL_ASSESSED_VALUE * ((1.0 + 0.05) ** years_from_fy26)
    if scenario == SCENARIO_SPILLOVER:
        haircut_amount = (
            VACANT_DC_ZONED_AV_FY26
            * ASSESSED_VALUE_HAIRCUT_CONTINGENT_OVERLAY
            * IMPAIRMENT_DISCLOSURE_PROBABILITY_24MO
        )
        base = FY26_TOTAL_ASSESSED_VALUE - haircut_amount
        # Grow at 2%/yr -- residential + non-DC commercial LOW (-1.5% res + 0% com)
        # partially offset by existing DC CE&P depreciation replacement.
        return base * ((1.0 + 0.02) ** years_from_fy26)
    if scenario == SCENARIO_PARTIAL:
        if fy <= 2028:
            haircut_amount = (
                VACANT_DC_ZONED_AV_FY26
                * ASSESSED_VALUE_HAIRCUT_CONTINGENT_OVERLAY
                * IMPAIRMENT_DISCLOSURE_PROBABILITY_24MO
            )
            base = FY26_TOTAL_ASSESSED_VALUE - haircut_amount
            return base * ((1.0 + 0.03) ** years_from_fy26)
        # FY29+: recover half the haircut; grow at 3.5%/yr.
        haircut_amount = (
            VACANT_DC_ZONED_AV_FY26
            * ASSESSED_VALUE_HAIRCUT_CONTINGENT_OVERLAY
            * IMPAIRMENT_DISCLOSURE_PROBABILITY_24MO
            * 0.5
        )
        base = FY26_TOTAL_ASSESSED_VALUE - haircut_amount
        return base * ((1.0 + 0.035) ** years_from_fy26)
    raise ValueError(f"unknown scenario {scenario}")


def required_re_tax_rate(
    scenario: str, fy: int, revenue_total: float, expenditure_total: float,
    real_estate_tax_revenue: float, assessed_value: float
) -> float:
    """Return the EFFECTIVE (gross-of-relief) rate per $100 AV required
    to close any gap between total expenditure and non-RE revenue.

    Formula per brief:  rate = (expenditure - non-RE revenue) / (AV/100).
    non-RE revenue = total revenue - real-estate-tax revenue.

    Note on nominal vs. effective:
    - The nominal FY26 RE tax rate is $0.906/$100 (PWC-REV-FY26-30 p.2).
    - Tax relief, senior exemptions, exonerations, and deferrals reduce
      billed revenue by about 18% relative to nominal rate x gross AV.
      The FY26 effective rate implied by the adopted budget is therefore
      ~$0.746/$100 ($1,025.92M net RE revenue / $137.56B AV x 100).
    - This formula yields the *effective* rate; multiply by ~1.214 to
      convert to a comparable nominal rate.  See the paired metric
      'required_re_tax_rate_nominal_per_100'.
    """
    non_re_revenue = revenue_total - real_estate_tax_revenue
    gap = expenditure_total - non_re_revenue
    if assessed_value <= 0:
        return float("nan")
    return gap / (assessed_value / 100.0)


# FY26 effective vs. nominal ratio: $0.906 nominal / $0.746 effective = 1.214.
# This reflects tax relief + exonerations + deferrals held to FY26 levels.
NOMINAL_TO_EFFECTIVE_RATIO: float = 0.906 / 0.746


def sum_real_estate_tax_revenue(revenue_df: pd.DataFrame, scenario: str, fy: int) -> float:
    """RE tax revenue = residential + non-DC commercial + DC-share of DC tax total.

    The DC total line in our model bundles DC real property + CE&P + F&F
    + fees.  For the "RE tax" carve-out used by the required-rate formula,
    we use residential + non-DC commercial only (i.e. the portion actually
    levied via the real-estate tax rate).  The DC real-property portion is
    approx 38% of DC total; however, it is still levied via the same RE
    rate, so we include it too.
    """
    # Residential + non-DC commercial real estate (Stage 4 non-DC lines):
    g = revenue_df[
        (revenue_df["scenario"] == scenario)
        & (revenue_df["fiscal_year"] == fy)
    ]
    non_dc_re = g[
        g["revenue_source"].isin(
            ["residential_real_estate", "non_dc_commercial_real_estate"]
        )
    ]["amount_usd"].sum()
    # DC real property approximation -- 38% of DC total (per Stage 1
    # TY2024: $144.2M / $293.7M = 49%; and the pipeline ramp shifts this to
    # ~38-45% as CE&P grows faster than real property).  We use 40%.
    dc_total = g[g["revenue_source"] == "data_center_tax_revenue"]["amount_usd"].sum()
    dc_real_property_share = 0.40
    dc_re = dc_total * dc_real_property_share
    return float(non_dc_re + dc_re)


def debt_service_metrics(
    debt_service: float, revenue_total: float
) -> Tuple[float, float, bool]:
    """Return (ratio, headroom_usd, breached)."""
    if revenue_total <= 0:
        return float("nan"), float("nan"), True
    ratio = debt_service / revenue_total
    headroom = DEBT_SERVICE_CAP_FRACTION * revenue_total - debt_service
    breached = ratio > DEBT_SERVICE_CAP_FRACTION
    return ratio, headroom, breached


# ---------------------------------------------------------------------------
# Scenario results integration.
# ---------------------------------------------------------------------------


@dataclass
class ScenarioRow:
    scenario: str
    fiscal_year: int
    metric: str
    value: float
    unit: str


def build_scenario_results(
    revenue_df: pd.DataFrame,
    exp_df: pd.DataFrame,
    debt_totals: Dict[int, float],
) -> pd.DataFrame:
    rev_tot = revenue_totals_and_split_base(revenue_df)
    exp_tot = expenditure_totals(exp_df)

    rows: List[ScenarioRow] = []

    # Reserve trajectory state per scenario.
    reserves: Dict[str, float] = {s: FY26_UNASSIGNED_GF_BALANCE for s in SCENARIOS}

    for scenario in SCENARIOS:
        for fy in FISCAL_YEARS:
            rev_row = rev_tot[(rev_tot["scenario"] == scenario) & (rev_tot["fiscal_year"] == fy)].iloc[0]
            exp_row = exp_tot[(exp_tot["scenario"] == scenario) & (exp_tot["fiscal_year"] == fy)].iloc[0]
            total_rev = float(rev_row["total_revenue"])
            split_base = float(rev_row["split_base_eligible"])
            total_exp = float(exp_row["total_committed_expenditure"])
            surplus = total_rev - total_exp

            rows.append(ScenarioRow(scenario, fy, "total_revenue", total_rev, "USD"))
            rows.append(
                ScenarioRow(scenario, fy, "split_base_eligible_for_schools_transfer",
                           split_base, "USD")
            )
            rows.append(
                ScenarioRow(scenario, fy, "schools_transfer",
                           schools_transfer(split_base), "USD")
            )
            rows.append(ScenarioRow(scenario, fy, "total_committed_expenditure", total_exp, "USD"))
            rows.append(ScenarioRow(scenario, fy, "surplus_or_deficit", surplus, "USD"))

            # Required RE rate to close the gap.
            re_rev = sum_real_estate_tax_revenue(revenue_df, scenario, fy)
            av = scenario_assessed_value(scenario, fy)
            required_rate = required_re_tax_rate(
                scenario, fy, total_rev, total_exp, re_rev, av
            )
            rows.append(ScenarioRow(scenario, fy, "real_estate_tax_revenue", re_rev, "USD"))
            rows.append(ScenarioRow(scenario, fy, "assessed_value_base", av, "USD"))
            rows.append(
                ScenarioRow(scenario, fy, "required_re_tax_rate_effective_per_100",
                           required_rate, "USD_per_100_AV")
            )
            required_rate_nominal = required_rate * NOMINAL_TO_EFFECTIVE_RATIO
            rows.append(
                ScenarioRow(scenario, fy, "required_re_tax_rate_nominal_per_100",
                           required_rate_nominal, "USD_per_100_AV")
            )
            rows.append(
                ScenarioRow(
                    scenario, fy, "required_re_rate_delta_vs_fy26_nominal",
                    required_rate_nominal - FY26_RE_TAX_RATE_PER_100, "USD_per_100_AV"
                )
            )

            # Debt-service metrics.
            ds = debt_totals[fy]
            ratio, headroom, breached = debt_service_metrics(ds, total_rev)
            rows.append(ScenarioRow(scenario, fy, "debt_service_total", ds, "USD"))
            rows.append(ScenarioRow(scenario, fy, "debt_service_to_revenue_ratio",
                                   ratio, "fraction"))
            rows.append(ScenarioRow(scenario, fy, "debt_service_headroom_vs_10pct_cap",
                                   headroom, "USD"))
            rows.append(ScenarioRow(scenario, fy, "debt_service_cap_breached",
                                   1.0 if breached else 0.0, "boolean"))

            # Reserve trajectory: unassigned GF balance = prior + surplus.
            reserves[scenario] = reserves[scenario] + surplus
            reserve_floor = RE_POLICY_UNASSIGNED_FLOOR_FRACTION * total_rev
            floor_breached = reserves[scenario] < reserve_floor
            rows.append(
                ScenarioRow(scenario, fy, "unassigned_gf_balance_eoy",
                           reserves[scenario], "USD")
            )
            rows.append(
                ScenarioRow(scenario, fy, "unassigned_gf_balance_policy_floor_7_5pct",
                           reserve_floor, "USD")
            )
            rows.append(
                ScenarioRow(scenario, fy, "unassigned_gf_balance_floor_breached",
                           1.0 if floor_breached else 0.0, "boolean")
            )

    return pd.DataFrame([r.__dict__ for r in rows])


SCENARIO_SPILLOVER_IMPAIRMENT_TRIGGERED: str = "CAPEX Spillover (impairment triggered)"


def impairment_sensitivity_results(
    revenue_df: pd.DataFrame,
    exp_df: pd.DataFrame,
    debt_totals: Dict[int, float],
) -> pd.DataFrame:
    """CAPEX Spillover one-time FY28 full impairment markdown sensitivity.

    If the 40% impairment-disclosure probability triggers, then the FY28
    assessed value takes the FULL 25% haircut (not the expected-value 10%
    = 25% x 40%).  This markdown propagates through:
      - DC tax revenue (one-time FY28 hit, partially reversing in FY29+ if
        re-assessments stabilize).
      - Assessed-value denominator in debt-to-AV and required-rate checks.
    """
    rev_tot = revenue_totals_and_split_base(revenue_df)
    exp_tot = expenditure_totals(exp_df)

    # Incremental markdown beyond what CAPEX Spillover already applied:
    # base has expected (25% x 40%) = 10% haircut.  A realized trigger
    # would take the full 25%; delta = 15% of VACANT_DC_ZONED_AV_FY26.
    extra_haircut = VACANT_DC_ZONED_AV_FY26 * ASSESSED_VALUE_HAIRCUT_CONTINGENT_OVERLAY * (
        1.0 - IMPAIRMENT_DISCLOSURE_PROBABILITY_24MO
    )
    extra_re_loss = extra_haircut * (FY26_RE_TAX_RATE_PER_100 / 100.0)

    rows: List[ScenarioRow] = []
    reserves_triggered = FY26_UNASSIGNED_GF_BALANCE
    tag = SCENARIO_SPILLOVER_IMPAIRMENT_TRIGGERED

    for fy in FISCAL_YEARS:
        rev_row = rev_tot[(rev_tot["scenario"] == SCENARIO_SPILLOVER) & (rev_tot["fiscal_year"] == fy)].iloc[0]
        exp_row = exp_tot[(exp_tot["scenario"] == SCENARIO_SPILLOVER) & (exp_tot["fiscal_year"] == fy)].iloc[0]
        total_rev = float(rev_row["total_revenue"])
        split_base = float(rev_row["split_base_eligible"])
        total_exp = float(exp_row["total_committed_expenditure"])

        # FY28 one-time hit; propagates forward at the same extra RE loss.
        if fy >= 2028:
            total_rev = total_rev - extra_re_loss
            split_base = split_base - extra_re_loss
            # Schools transfer re-solves on the shocked base.
            new_schools = schools_transfer(split_base)
            # Expenditure total changes by (new_schools - old_schools).
            old_schools = float(
                exp_df[
                    (exp_df["scenario"] == SCENARIO_SPILLOVER)
                    & (exp_df["fiscal_year"] == fy)
                    & (exp_df["block"] == "schools_transfer")
                ]["amount_usd"].iloc[0]
            )
            total_exp = total_exp - old_schools + new_schools

        surplus = total_rev - total_exp
        reserves_triggered = reserves_triggered + surplus

        rows.append(ScenarioRow(tag, fy, "total_revenue", total_rev, "USD"))
        rows.append(ScenarioRow(tag, fy, "split_base_eligible_for_schools_transfer",
                               split_base, "USD"))
        rows.append(ScenarioRow(tag, fy, "schools_transfer",
                               schools_transfer(split_base), "USD"))
        rows.append(ScenarioRow(tag, fy, "total_committed_expenditure",
                               total_exp, "USD"))
        rows.append(ScenarioRow(tag, fy, "surplus_or_deficit",
                               surplus, "USD"))
        rows.append(ScenarioRow(tag, fy, "unassigned_gf_balance_eoy",
                               reserves_triggered, "USD"))
        rows.append(ScenarioRow(tag, fy,
                               "unassigned_gf_balance_policy_floor_7_5pct",
                               RE_POLICY_UNASSIGNED_FLOOR_FRACTION * total_rev, "USD"))
        rows.append(ScenarioRow(tag, fy,
                               "unassigned_gf_balance_floor_breached",
                               1.0 if reserves_triggered < RE_POLICY_UNASSIGNED_FLOOR_FRACTION * total_rev else 0.0,
                               "boolean"))
        ds = debt_totals[fy]
        ratio, headroom, breached = debt_service_metrics(ds, total_rev)
        rows.append(ScenarioRow(tag, fy, "debt_service_total", ds, "USD"))
        rows.append(ScenarioRow(tag, fy,
                               "debt_service_to_revenue_ratio", ratio, "fraction"))
        rows.append(ScenarioRow(tag, fy,
                               "debt_service_headroom_vs_10pct_cap", headroom, "USD"))
        rows.append(ScenarioRow(tag, fy,
                               "debt_service_cap_breached", 1.0 if breached else 0.0, "boolean"))

    return pd.DataFrame([r.__dict__ for r in rows])


# ---------------------------------------------------------------------------
# Validation: compare Pre-Cancellation Digital Gateway against the published
# 5-year forecast.
# ---------------------------------------------------------------------------


def validate_pre_dg(rev_tot: pd.DataFrame) -> pd.DataFrame:
    """For each FY in FY27-FY30 (the published range), compare the
    Pre-Cancellation Digital Gateway split-base to the published GR
    split-base. Warn if |delta| > 3%."""
    rows = []
    for fy in (2027, 2028, 2029, 2030):
        model_split = float(
            rev_tot[(rev_tot["scenario"] == SCENARIO_PRE_DG) & (rev_tot["fiscal_year"] == fy)][
                "split_base_eligible"
            ].iloc[0]
        )
        published = PUBLISHED_GR_BASE[fy]
        delta = model_split - published
        delta_pct = delta / published
        rows.append(
            {
                "fiscal_year": fy,
                "model_pre_cancellation_dg_split_base": model_split,
                "published_FY26_FY30_forecast": published,
                "delta_usd": delta,
                "delta_pct": delta_pct,
                "within_3pct": abs(delta_pct) <= 0.03,
            }
        )
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# CSV and XLSX writers.
# ---------------------------------------------------------------------------


def write_scenario_results_csv(df: pd.DataFrame, path: Path) -> None:
    df.to_csv(path, index=False)


def write_revenue_detail_csv(df: pd.DataFrame, path: Path) -> None:
    # Long-format: scenario, fiscal_year, revenue_source, amount_usd, in_split_base.
    df.to_csv(path, index=False)


def write_expenditure_detail_csv(df: pd.DataFrame, path: Path) -> None:
    df.to_csv(path, index=False)


def write_xlsx(
    revenue_df: pd.DataFrame,
    exp_df: pd.DataFrame,
    scenario_results: pd.DataFrame,
    impairment_results: pd.DataFrame,
    validation_df: pd.DataFrame,
    path: Path,
) -> None:
    """Build a shareable workbook.  One sheet per scenario plus summary
    and assumptions sheets.  Uses openpyxl to avoid a pandas ExcelWriter
    pinned-version dependency."""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required for XLSX output.  Install with "
            "`pip install --break-system-packages openpyxl`."
        ) from e

    wb = Workbook()
    # Remove default sheet.
    wb.remove(wb.active)

    hdr_font = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")

    def _write_df(ws, df: pd.DataFrame) -> None:
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=j, value=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for i, (_, r) in enumerate(df.iterrows(), start=2):
            for j, col in enumerate(df.columns, start=1):
                v = r[col]
                ws.cell(row=i, column=j, value=v)
        # Set column widths modestly.
        for j, col in enumerate(df.columns, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = max(
                12, min(60, len(str(col)) + 4)
            )

    # --- Summary / comparison sheet ---
    ws = wb.create_sheet("summary")
    pivot_metrics = [
        "total_revenue",
        "total_committed_expenditure",
        "surplus_or_deficit",
        "schools_transfer",
        "required_re_tax_rate_nominal_per_100",
        "required_re_tax_rate_effective_per_100",
        "debt_service_to_revenue_ratio",
        "unassigned_gf_balance_eoy",
    ]
    summary_rows = []
    for metric in pivot_metrics:
        for scenario in SCENARIOS:
            d = scenario_results[
                (scenario_results["scenario"] == scenario)
                & (scenario_results["metric"] == metric)
            ].set_index("fiscal_year")["value"]
            row = {"scenario": scenario, "metric": metric}
            for fy in FISCAL_YEARS:
                row[f"FY{fy}"] = float(d.loc[fy]) if fy in d.index else None
            summary_rows.append(row)
    _write_df(ws, pd.DataFrame(summary_rows))

    # --- Per-scenario sheet ---
    for scenario in SCENARIOS:
        # openpyxl tab names must be <=31 chars; use the short form.
        tab_name = SHORT_NAMES.get(scenario, scenario)
        ws = wb.create_sheet(tab_name)
        # Year-columns matrix of all metrics for the scenario.
        d = scenario_results[scenario_results["scenario"] == scenario]
        matrix_rows = []
        for metric in d["metric"].unique():
            row = {"metric": metric}
            sub = d[d["metric"] == metric].set_index("fiscal_year")["value"]
            for fy in FISCAL_YEARS:
                row[f"FY{fy}"] = float(sub.loc[fy]) if fy in sub.index else None
            matrix_rows.append(row)
        _write_df(ws, pd.DataFrame(matrix_rows))

    # --- Revenue detail sheet ---
    ws = wb.create_sheet("revenue_detail")
    pivot_rev = revenue_df.pivot_table(
        index=["scenario", "revenue_source"],
        columns="fiscal_year",
        values="amount_usd",
        aggfunc="sum",
    ).reset_index()
    _write_df(ws, pivot_rev)

    # --- Expenditure detail sheet ---
    ws = wb.create_sheet("expenditure_detail")
    pivot_exp = exp_df.pivot_table(
        index=["scenario", "block", "line_item"],
        columns="fiscal_year",
        values="amount_usd",
        aggfunc="sum",
    ).reset_index()
    _write_df(ws, pivot_exp)

    # --- Impairment sensitivity sheet ---
    ws = wb.create_sheet("impairment_sensitivity")
    pivot_imp = impairment_results.pivot_table(
        index="metric",
        columns="fiscal_year",
        values="value",
        aggfunc="sum",
    ).reset_index()
    _write_df(ws, pivot_imp)

    # --- Validation sheet ---
    ws = wb.create_sheet("validation_pre_cancellation_dg")
    _write_df(ws, validation_df)

    # --- Assumptions sheet: dumps every parameter CSV we read ---
    ws = wb.create_sheet("assumptions")
    compiled = []
    for name, fname in [
        ("spillover_parameters", "spillover_parameters.csv"),
        ("revenue_driver_assumptions", "revenue_driver_assumptions.csv"),
        ("expenditure_assumptions", "expenditure_assumptions.csv"),
    ]:
        df = _read_csv_skip_comments(DATA / fname)
        df.insert(0, "source_csv", name)
        compiled.append(df)
    # All CSVs have slightly different column layouts; concatenate with
    # union of columns.
    merged = pd.concat(compiled, ignore_index=True, sort=False)
    _write_df(ws, merged)

    # --- README sheet ---
    ws = wb.create_sheet("README")
    readme_lines = [
        ["PWC Fiscal Risk Model, Integrated Scenarios FY27-FY31"],
        [""],
        ["Pre-Cancellation Digital Gateway."],
        [
            "Digital Gateway proceeds on pre-ruling schedule. Uses published FY26-FY30 "
            "adopted 5-year forecast for the GR split base (PWC-REV-FY26-30 p.2)."
        ],
        ["CAPEX Spillover."],
        [
            "Canceled Pageland projects removed; new non-Pageland overlay CAPEX "
            "deterred at 35/25/18/12/10% year-1..year-5; contingent overlay assessed "
            "value marks down at expected-value 25% x 40% = 10%."
        ],
        ["Partial Recovery."],
        [
            "Canceled Pageland stays canceled; half of deterred CAPEX returns "
            "from FY29 at 70% nominal yield; impairment probability drops to 15%."
        ],
        [""],
        ["Reproducibility: run `python3 model/pwc_5yr.py`."],
        [""],
        ["Authoritative equations:"],
        ["  Required RE rate = (expenditure - non-RE revenue) / (assessed value / 100)"],
        ["  Schools transfer = 57.23% * eligible revenue base (2013 Agreement)"],
        ["  Debt service cap = 10% of revenues (PSFM 5.02(d))"],
        ["  Unassigned GF floor = 7.5% of GF revenues (PSFM 1.02)"],
    ]
    for i, line in enumerate(readme_lines, start=1):
        for j, v in enumerate(line, start=1):
            c = ws.cell(row=i, column=j, value=v)
            if i in (1, 3, 5, 7) and j == 1:
                c.font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 100

    wb.save(path)


# ---------------------------------------------------------------------------
# Main.
# ---------------------------------------------------------------------------


def main() -> int:
    print(f"Running Stage 6 integrated model from {REPO}")

    # Build revenue detail.
    revenue_df = build_revenue_rows()

    # Build expenditure detail, with Schools transfer re-solved per-scenario.
    rev_tot = revenue_totals_and_split_base(revenue_df)
    exp_df = build_expenditure_rows(rev_tot)

    # Debt-service totals from Stage 5 CSV.
    debt_totals = load_debt_schedule_totals()

    # Scenario-level metrics.
    scenario_results = build_scenario_results(revenue_df, exp_df, debt_totals)

    # Impairment sensitivity.
    impairment_results = impairment_sensitivity_results(revenue_df, exp_df, debt_totals)

    # Validation vs. published FY26-FY30 forecast.
    validation = validate_pre_dg(rev_tot)

    # Write outputs.
    write_revenue_detail_csv(revenue_df, DATA / "scenario_revenue_detail.csv")
    write_expenditure_detail_csv(exp_df, DATA / "scenario_expenditure_detail.csv")
    write_scenario_results_csv(
        pd.concat([scenario_results, impairment_results], ignore_index=True),
        DATA / "scenario_results.csv",
    )
    write_xlsx(
        revenue_df,
        exp_df,
        scenario_results,
        impairment_results,
        validation,
        MODEL / "pwc_5yr.xlsx",
    )

    # Terminal report.
    print("\n=== Validation: Pre-Cancellation Digital Gateway vs. published FY26-FY30 forecast ===")
    print(validation.to_string(index=False))

    print("\n=== Headline surplus/deficit by scenario (USD millions) ===")
    for scenario in SCENARIOS:
        print(f"  {scenario}:")
        for fy in FISCAL_YEARS:
            v = float(
                scenario_results[
                    (scenario_results["scenario"] == scenario)
                    & (scenario_results["fiscal_year"] == fy)
                    & (scenario_results["metric"] == "surplus_or_deficit")
                ]["value"].iloc[0]
            )
            print(f"    FY{fy}: ${v/1e6:+,.1f}M")

    print("\n=== Required RE tax rate per $100 (CAPEX Spillover) ===")
    print("  (Nominal rate; current FY26 nominal = $0.906; effective = ~$0.746)")
    for fy in FISCAL_YEARS:
        v = float(
            scenario_results[
                (scenario_results["scenario"] == SCENARIO_SPILLOVER)
                & (scenario_results["fiscal_year"] == fy)
                & (scenario_results["metric"] == "required_re_tax_rate_nominal_per_100")
            ]["value"].iloc[0]
        )
        print(f"  FY{fy}: ${v:.4f}/$100 (current FY26 ${FY26_RE_TAX_RATE_PER_100:.3f})")

    print("\n=== Debt-service cap breaches ===")
    for scenario in SCENARIOS:
        flags = [
            int(
                scenario_results[
                    (scenario_results["scenario"] == scenario)
                    & (scenario_results["fiscal_year"] == fy)
                    & (scenario_results["metric"] == "debt_service_cap_breached")
                ]["value"].iloc[0]
            )
            for fy in FISCAL_YEARS
        ]
        first_breach = None
        for fy, f in zip(FISCAL_YEARS, flags):
            if f:
                first_breach = fy
                break
        print(f"  {scenario}: first breach = {first_breach}")

    print("\n=== Unassigned GF balance 7.5% floor breaches ===")
    for scenario in SCENARIOS:
        flags = [
            int(
                scenario_results[
                    (scenario_results["scenario"] == scenario)
                    & (scenario_results["fiscal_year"] == fy)
                    & (scenario_results["metric"] == "unassigned_gf_balance_floor_breached")
                ]["value"].iloc[0]
            )
            for fy in FISCAL_YEARS
        ]
        first_breach = None
        for fy, f in zip(FISCAL_YEARS, flags):
            if f:
                first_breach = fy
                break
        print(f"  {scenario}: first breach = {first_breach}")

    print("\nWrote:")
    print(f"  {DATA / 'scenario_results.csv'}")
    print(f"  {DATA / 'scenario_revenue_detail.csv'}")
    print(f"  {DATA / 'scenario_expenditure_detail.csv'}")
    print(f"  {MODEL / 'pwc_5yr.xlsx'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
